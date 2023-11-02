#!/usr/bin/python3

import math 
from openpyxl import load_workbook 

wb = load_workbook(filename = 'student.xlsx')
ws = wb.active
total_list = []

for row in ws.iter_rows(min_row = 1, max_row = ws.max_row):
	if type(row[2].value) != str and type(row[3].value) != str and type(row[4].value) != str and type(row[5].value) != str:
		total = float(row[2].value) * 0.3 + float(row[3].value) * 0.35 + float(row[4].value) * 0.34 + float(row[5].value) * 0.01
		ws.cell(row = row[0].row, column = 7, value = total)
		total_list.append(total)	
 
A_students = math.trunc((ws.max_row - 2 + 1) * 0.3)
B_students = math.trunc((ws.max_row - 2 + 1) * 0.7)
A_list = []
B_list = []
C_list = []
A_index_list = []
B_index_list = []
C_index_list = []

cnt = 0 
while cnt < A_students:
	A_max = 0
	A_index = -1
	for index, t in enumerate(total_list):
		if index not in A_index_list:
			if A_max <= t:
				A_max = t
				A_index = index
	if A_max < 40:
		break
	
	if cnt == A_students -1 and A_max in A_list:
		for i, l in enumerate(A_list):
			if l == A_max:
				A_list.remove(l)
				del A_index_list[i]
		break

	A_list.append(A_max)
	A_index_list.append(A_index)
	cnt += 1
    
cnt = 0
while cnt < B_students - A_students:
	B_max = 0
	B_index = -1
	for index, t in enumerate(total_list):
		if index not in A_index_list:
			if index not in B_index_list:
				if B_max <= t:
					B_max = t
					B_index = index
	if B_max < 40:
		break

	if cnt == B_students -1 and B_max in B_list:
		for i, l in enumerate(B_list):
			if l == B_max:
				B_list.remove(l)
				del B_index_list[i]
		break

	B_list.append(B_max)	
	B_index_list.append(B_index)
	cnt += 1
 
while 1==1:
	C_max = 0
	C_index = -1
	for index, t in enumerate(total_list):
		if index not in A_index_list:
			if index not in B_index_list:
				if index not in C_index_list:
					if C_max <= t:
						C_max = t
						C_index = index
	if C_max < 40:
		break
	C_list.append(C_max)
	C_index_list.append(C_index)
 
for index, a in enumerate(A_index_list):
	if index < math.trunc(len(A_index_list)/2):
		if index == math.trunc(len(A_index_list)/2) - 1 and A_list[index] == A_list[index + 1]:
			for i, d in enumerate(A_list):
				if d == A_list[index]:
					ws.cell(row = A_index_list[i] + 2, column = 8, value = 'A0')
		else:
			ws.cell(row = a + 2, column = 8, value = 'A+')
	else:
		ws.cell(row = a + 2, column = 8, value = 'A0')

for index, b in enumerate(B_index_list):
	if index < math.trunc(len(B_index_list)/2):
		if index == math.trunc(len(B_index_list)/2) - 1 and B_list[index] == B_list[index + 1]:
			for i, d in enumerate(B_list):
				if d == B_list[index]:
					ws.cell(row = B_index_list[i] + 2, column = 8, value = 'B0')
		else:
			ws.cell(row = b + 2, column = 8, value = 'B+')
	else:
		ws.cell(row = b + 2, column = 8, value = 'B0')

for index, c in enumerate(C_index_list):
	if index < math.trunc(len(C_index_list)/2):
		if index == math.trunc(len(C_index_list)/2) - 1 and C_list[index] == C_list[index + 1]:
			for i, d in enumerate(C_list):
				if d == C_list[index]:
					ws.cell(row = C_index_list[i] + 2, column = 8, value = 'C0')
		else:
			ws.cell(row = c + 2, column = 8, value = 'C+')
	else:
		ws.cell(row = c + 2, column = 8, value = 'C0')

for t in total_list:
	if t < 40:
		ws.cell(row = total_list.index(t) + 2, column = 8, value = 'F')
  
wb.save('student.xlsx')